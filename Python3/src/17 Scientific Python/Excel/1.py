import openpyxl
mywb = openpyxl.load_workbook('w.xlsx')

print(mywb.sheetnames)

mywb.create_sheet()
print(mywb.sheetnames)

mywb.create_sheet(index=0, title='1st Sheet')
print(mywb.sheetnames)


mywb.create_sheet(index=2, title='2nd Sheet')
print(mywb.sheetnames)



mywb.save('example_filetest.xlsx')